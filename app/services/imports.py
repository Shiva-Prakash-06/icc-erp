"""Staged, auditable imports for the supplied ICC/OIA source material."""

from __future__ import annotations

import hashlib
import csv
import io
import re
from datetime import date, datetime, time, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from docx import Document as WordDocument
from openpyxl import Workbook
from werkzeug.utils import secure_filename

from app.database import db
from app.models.erp import (
    ChecklistInstance,
    ChecklistItemStatus,
    ChecklistTemplate,
    ChecklistTemplateItem,
    DocumentRecord,
    ImportBatch,
    ImportRow,
    OperatingUnit,
    Person,
    ProjectSession,
    ReportSnapshot,
    SessionAttendance,
    TeamAssignment,
    Wing,
    WorkTask,
)
from app.models.project import AcademicYear, Campus, ProgramType, Project, ProjectParticipant
from app.models.production import ControlledVocabulary, Position, ReportDefinition
from app.services.audit import record_audit
from app.services.drive import validate_drive_link


IST = ZoneInfo("Asia/Kolkata")
ROOT = Path(__file__).resolve().parents[3]
SOURCE_PATHS = {
    "events_summary": ROOT / "2026 ICC EVENTS REPORT SUMMARY.xlsx",
    "coffee_meet": ROOT / "COFFEE MEET & GREET",
    "summer_school": ROOT / "_Summer School- Check List.xlsx",
}
IMPORT_SCHEMA_VERSION = 2

STANDARD_IMPORTS = {
    "people": (["registration_number", "first_name", "last_name", "email", "phone", "campus_code", "person_type", "nationality"], ["first_name"]),
    "icc_roster": (["registration_number", "email", "wing_code", "role_label", "academic_year"], ["wing_code", "role_label", "academic_year"]),
    "projects": (["code", "title", "campus_code", "program_type", "academic_year", "start_date", "end_date", "project_type", "category", "unit_code", "wing_code"], ["code", "title", "campus_code", "program_type", "academic_year", "start_date", "end_date"]),
    "participants": (["project_code", "registration_number", "email", "participant_type"], ["project_code", "participant_type"]),
    "attendance": (["project_code", "session_code", "registration_number", "email", "status"], ["project_code", "session_code", "status"]),
    "checklists": (["project_code", "template_code", "item_code", "title", "category", "mandatory", "owner"], ["project_code", "template_code", "item_code", "title"]),
    "documents": (["project_code", "title", "category", "status", "drive_url", "classification", "mandatory_for_closure"], ["project_code", "title", "category"]),
}


def _json_value(value):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def _sha256_path(path: Path):
    digest = hashlib.sha256()
    if path.is_dir():
        files = sorted(item for item in path.rglob("*") if item.is_file())
    else:
        files = [path]
    for item in files:
        digest.update(str(item.relative_to(path) if path.is_dir() else item.name).encode())
        with item.open("rb") as handle:
            for block in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(block)
    return digest.hexdigest()


def _new_batch(import_type, source):
    source = Path(source).resolve()
    digest = _sha256_path(source)
    key = f"v{IMPORT_SCHEMA_VERSION}:{import_type}:{digest}"
    existing = ImportBatch.query.filter_by(idempotency_key=key).first()
    if existing:
        return existing, False
    batch = ImportBatch(
        idempotency_key=key,
        import_type=import_type,
        source_file=str(source),
        source_sha256=digest,
        status="Staged",
    )
    db.session.add(batch)
    db.session.flush()
    return batch, True


def _add_row(batch, sheet, row_number, source, normalized, entity, messages=None):
    messages = messages or []
    row = ImportRow(
        batch_id=batch.id,
        sheet_name=sheet,
        source_row=row_number,
        source_json={key: _json_value(value) for key, value in source.items()},
        normalized_json=normalized,
        validation_status="Error" if messages else "Valid",
        validation_messages=messages,
        target_entity=entity,
    )
    db.session.add(row)
    batch.staged_count += 1
    if messages:
        batch.error_count += 1
    else:
        batch.valid_count += 1
    return row


def stage_supplied_source(import_type):
    if import_type not in SOURCE_PATHS:
        raise ValueError("Unknown supplied source type.")
    source = SOURCE_PATHS[import_type]
    if not source.exists():
        raise FileNotFoundError(source)
    batch, created = _new_batch(import_type, source)
    if not created:
        return batch
    if import_type == "events_summary":
        _stage_events(batch, source)
    elif import_type == "coffee_meet":
        _stage_coffee(batch, source)
    else:
        _stage_summer_school(batch, source)
    record_audit("import.stage", batch, after={"type": import_type, "rows": batch.staged_count})
    db.session.commit()
    return batch


def build_import_template(import_type):
    if import_type not in STANDARD_IMPORTS:
        raise ValueError("Unknown standard import type.")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = import_type
    columns, _ = STANDARD_IMPORTS[import_type]
    sheet.append(columns)
    sheet.freeze_panes = "A2"
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def stage_uploaded_source(import_type, uploaded_file, operator_key):
    if import_type not in STANDARD_IMPORTS:
        raise ValueError("Unknown standard import type.")
    filename = secure_filename(uploaded_file.filename or "")
    suffix = Path(filename).suffix.lower()
    if suffix not in {".xlsx", ".csv"}:
        raise ValueError("Only .xlsx and .csv import files are accepted; macro-enabled files are rejected.")
    content = uploaded_file.read()
    digest = hashlib.sha256(content).hexdigest()
    key = f"v{IMPORT_SCHEMA_VERSION}:{import_type}:{operator_key}:{digest}"
    existing = ImportBatch.query.filter_by(idempotency_key=key).first()
    if existing:
        return existing
    batch = ImportBatch(idempotency_key=key, import_type=import_type, source_file=filename, source_sha256=digest, status="Staged", importer_version=str(IMPORT_SCHEMA_VERSION), mapping_version="1")
    db.session.add(batch)
    db.session.flush()
    columns, required = STANDARD_IMPORTS[import_type]
    if suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip().lower() for value in next(values)]
        rows = ((sheet.title, number, dict(zip(headers, row))) for number, row in enumerate(values, start=2))
    else:
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        rows = ((filename, number, {str(key).strip().lower(): value for key, value in row.items()}) for number, row in enumerate(reader, start=2))
    available_headers = headers if suffix == ".xlsx" else [str(value or "").strip().lower() for value in (reader.fieldnames or [])]
    missing_headers = [column for column in required if column not in available_headers]
    if missing_headers:
        raise ValueError("Missing required columns: " + ", ".join(missing_headers))
    for sheet_name, row_number, source in rows:
        if not any(value not in (None, "") for value in source.values()):
            continue
        normalized = {column: _json_value(source.get(column)) for column in columns}
        messages = [f"{column} is required." for column in required if normalized.get(column) in (None, "")]
        if import_type in {"people", "participants", "attendance", "icc_roster"} and not (normalized.get("registration_number") or normalized.get("email")):
            messages.append("registration_number or email is required.")
        _add_row(batch, sheet_name, row_number, source, normalized, import_type, messages)
    record_audit("import.stage", batch, after={"type": import_type, "rows": batch.staged_count, "source_sha256": digest})
    db.session.commit()
    return batch


def _stage_events(batch, path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["2026 ICC EVENTS"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        source = dict(zip(headers, values))
        title = (source.get("EVENT NAME") or "").strip()
        if not title:
            continue
        normalized = {
            "title": title,
            "event_date": _json_value(source.get("DATE")),
            "time_range": source.get("EVENT TIME"),
            "venue": source.get("VENUE"),
            "target_audience": source.get("PARTICIPANTS"),
            "documents": [
                {"category": label, "url": source.get(column)}
                for column, label in [
                    ("EVENT POSTER", "Event Poster"),
                    ("STAGE BACKDROP POSTER", "Stage Backdrop"),
                    ("PROGRAM DETAILS", "Program Details"),
                    ("IMAGES/PHOTOS/FLICR", "Images/Photos"),
                    ("EVENT REPORT", "Event Report"),
                ]
                if source.get(column)
            ],
        }
        messages = [] if normalized["event_date"] else ["Event date is required."]
        _add_row(batch, sheet.title, row_number, source, normalized, "Project", messages)


def _stage_coffee(batch, folder):
    workbook_path = next(folder.glob("Final Volunteer List*.xlsx"))
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    roster = workbook["STU LIST"]
    current_role = None
    section = None
    for row_number, values in enumerate(roster.iter_rows(min_row=3, values_only=True), start=3):
        serial, role, name, programme, registration, student_type, attendance = values[:7]
        if isinstance(serial, str) and serial.strip():
            section = serial.strip()
            continue
        if not isinstance(serial, (int, float)) or not name:
            continue
        if role:
            current_role = str(role).strip()
        normalized = {
            "name": str(name).strip(),
            "registration_number": str(int(registration)) if isinstance(registration, float) else (str(registration).strip() if registration else None),
            "programme": str(programme).strip() if programme else None,
            "category": str(student_type).strip() if student_type else None,
            "role": current_role or ("Volunteer" if section == "Volunteers" else section),
            "section": section,
            "attendance": str(attendance).strip() if attendance else None,
        }
        _add_row(
            batch,
            "STU LIST",
            row_number,
            {"serial": serial, "role": role, "name": name, "programme": programme, "registration": registration, "student_type": student_type, "attendance": attendance},
            normalized,
            "Person",
            [] if normalized["name"] else ["Name is required."],
        )

    actions = workbook["Sheet1"]
    for row_number, values in enumerate(actions.iter_rows(min_row=3, values_only=True), start=3):
        serial, title, details = values[:3]
        if not title:
            continue
        _add_row(
            batch,
            "Sheet1",
            row_number,
            {"serial": serial, "title": title, "details": details},
            {"title": str(title).strip(), "details": str(details).strip() if details else None},
            "WorkTask",
        )

    document_row = 1
    for file in sorted(folder.iterdir()):
        if not file.is_file() or file == workbook_path:
            continue
        _add_row(
            batch,
            "FOLDER INDEX",
            document_row,
            {"file_name": file.name, "file_size": file.stat().st_size},
            {"title": file.stem, "file_name": file.name, "extension": file.suffix.lower()},
            "DocumentRecord",
        )
        document_row += 1

    schedule = [
        ("15:10", "15:15", "Welcome Performance"),
        ("15:15", "15:20", "Welcome Address and Inspirational Message"),
        ("15:20", "15:25", "Cultural Dance Performance"),
        ("15:25", "15:35", "Introduction to OIA & ICC"),
        ("15:35", "15:40", "Student Testimonial"),
        ("15:40", "15:45", "Musical Performance"),
        ("15:45", "16:15", "Campus Tour"),
        ("16:15", "16:30", "Refreshments"),
    ]
    for sequence, (starts, ends, title) in enumerate(schedule, start=1):
        _add_row(
            batch,
            "CMG _26 Programme Schedule.jpg",
            sequence,
            {"starts": starts, "ends": ends, "title": title},
            {"code": f"CMG-{sequence:02d}", "starts": starts, "ends": ends, "title": title},
            "ProjectSession",
        )

    report_path = folder / "Event Report_ICC COFFEE MEET & GREET 2026.docx"
    report_document = WordDocument(report_path)
    paragraphs = [paragraph.text.strip() for paragraph in report_document.paragraphs if paragraph.text.strip()]
    narrative = "\n\n".join(paragraphs[1:])
    _add_row(
        batch,
        report_path.name,
        1,
        {"title": paragraphs[0] if paragraphs else report_path.stem, "paragraph_count": len(paragraphs)},
        {
            "title": paragraphs[0] if paragraphs else report_path.stem,
            "narrative": narrative,
            "actual_reach": 175,
            "audience_breakdown": {
                "international_and_oci_students": 75,
                "parents": 85,
                "other_guests_and_team_within_total": 15,
            },
        },
        "ReportSnapshot",
    )


def _stage_summer_school(batch, path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    section = "OIA Operations"
    section_sequence = 0
    for row_number, values in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        serial, title, owner, source_status = values[1:5]
        if serial == "S. No":
            section = "IGP Team"
            section_sequence = 0
            continue
        if not title:
            continue
        section_sequence += 1
        sensitive = bool(re.search(r"passport|visa|c-?form", str(title), flags=re.I))
        normalized = {
            "code": f"{'OIA' if section == 'OIA Operations' else 'IGP'}-{section_sequence:02d}",
            "title": str(title).strip(),
            "category": section,
            "external_owner": str(owner).strip() if owner else None,
            "source_status": str(source_status).strip() if source_status else None,
            "sensitive": sensitive,
            "mandatory": True,
        }
        _add_row(
            batch,
            sheet.title,
            row_number,
            {"serial": serial, "title": title, "owner": owner, "status": source_status},
            normalized,
            "ChecklistTemplateItem",
        )


def _reference_data():
    year = AcademicYear.query.filter_by(name="2026-2027").first()
    if not year:
        year = AcademicYear(name="2026-2027", start_date=date(2026, 6, 1), end_date=date(2027, 5, 31), is_current=True)
        db.session.add(year)
    campus = Campus.query.filter_by(code="CEN").first() or Campus.query.filter_by(name="Bangalore Central Campus").first()
    if not campus:
        campus = Campus(code="CEN", name="Bangalore Central Campus")
        db.session.add(campus)
    elif not campus.code:
        campus.code = "CEN"
    icc_type = ProgramType.query.filter_by(name="ICC").first()
    igp_type = ProgramType.query.filter_by(name="IGP").first()
    if not icc_type:
        icc_type = ProgramType(name="ICC")
        db.session.add(icc_type)
    if not igp_type:
        igp_type = ProgramType(name="IGP")
        db.session.add(igp_type)
    icc = OperatingUnit.query.filter_by(code="ICC").first()
    igp = OperatingUnit.query.filter_by(code="IGP").first()
    if not icc:
        icc = OperatingUnit(code="ICC", name="International Christite Community")
        db.session.add(icc)
    if not igp:
        igp = OperatingUnit(code="IGP", name="India Gateway Program")
        db.session.add(igp)
    db.session.flush()
    wings = {}
    for code, name in (("EVENTS", "Events"), ("MEDIA", "Media"), ("CULTURALS", "Culturals")):
        wing = Wing.query.filter_by(operating_unit_id=icc.id, code=code).first()
        if not wing:
            wing = Wing(operating_unit_id=icc.id, code=code, name=name)
            db.session.add(wing)
        wings[code] = wing
    db.session.flush()
    for code, name, unit, wing in (
        ("ICC_SECRETARY_USC", "ICC Secretary / USC", icc, None),
        ("ICC_EVENTS_HEAD", "ICC Events Head", icc, wings["EVENTS"]),
        ("ICC_MEDIA_HEAD", "ICC Media Head", icc, wings["MEDIA"]),
        ("ICC_CULTURALS_HEAD", "ICC Culturals Head", icc, wings["CULTURALS"]),
        ("IGP_HEAD", "IGP Head", igp, None),
    ):
        if not Position.query.filter_by(code=code).first():
            db.session.add(Position(code=code, name=name, operating_unit_id=unit.id, wing_id=getattr(wing, "id", None)))
    vocabulary = {
        "project_status": ["Draft", "Pending Approval", "Planned", "Active", "Closing", "Completed", "Cancelled", "Archived"],
        "task_status": ["Not Started", "In Progress", "Blocked", "Submitted", "Approved", "Rejected", "Waived", "Completed"],
        "document_classification": ["Public", "Internal", "Restricted"],
        "attendance_status": ["Present", "Absent", "Excused", "Late"],
    }
    for domain, values in vocabulary.items():
        for sort_order, label in enumerate(values):
            code = re.sub(r"[^A-Z0-9]+", "_", label.upper()).strip("_")
            if not ControlledVocabulary.query.filter_by(domain=domain, code=code, version=1).first():
                db.session.add(ControlledVocabulary(domain=domain, code=code, label=label, version=1, sort_order=sort_order))
    if not ReportDefinition.query.filter_by(code="PROJECT_OPERATIONAL", version=1).first():
        db.session.add(ReportDefinition(code="PROJECT_OPERATIONAL", name="Project Operational Report", report_type="Project Operational Report", version=1, schema_json={"required": ["project", "reach", "execution", "budget"]}))
    db.session.flush()
    return year, campus, icc_type, igp_type, icc, igp, wings["EVENTS"]


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_time_range(text):
    matches = re.findall(r"(\d{1,2})(?::(\d{2}))?\s*(am|pm)?", (text or "").lower())
    parsed = []
    for hour, minute, marker in matches[:2]:
        hour = int(hour)
        minute = int(minute or 0)
        if marker == "pm" and hour < 12:
            hour += 12
        if marker == "am" and hour == 12:
            hour = 0
        parsed.append(time(hour, minute))
    return parsed if len(parsed) == 2 else [time(9), time(17)]


def _code(prefix, sequence):
    return f"{prefix}-2026-CEN-{sequence:03d}"


def _as_bool(value):
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "required"}


def _person_for_import(data):
    registration = str(data.get("registration_number") or "").strip() or None
    email = str(data.get("email") or "").strip().lower() or None
    person = Person.query.filter_by(registration_number=registration).first() if registration else None
    if not person and email:
        person = Person.query.filter(db.func.lower(Person.primary_email) == email).first()
    return person


def _commit_standard_rows(batch, valid_rows, defaults):
    year, campus, icc_type, igp_type, icc, igp, events_wing = defaults
    committed = 0
    for sequence, row in enumerate(valid_rows, start=1):
        data = row.normalized_json
        target = None
        if batch.import_type == "people":
            target = _person_for_import(data)
            target_campus = Campus.query.filter_by(code=data.get("campus_code")).first() or campus
            if not target:
                target = Person(registration_number=data.get("registration_number") or None, first_name=data["first_name"], last_name=data.get("last_name") or None, primary_email=(data.get("email") or None), phone=data.get("phone") or None, campus_id=target_campus.id, person_type=data.get("person_type") or "Student", nationality_country=data.get("nationality") or None)
                db.session.add(target)
            else:
                target.first_name = data.get("first_name") or target.first_name
                target.last_name = data.get("last_name") or target.last_name
                target.phone = data.get("phone") or target.phone
        elif batch.import_type == "projects":
            target = Project.query.filter_by(code=data["code"]).first()
            target_campus = Campus.query.filter_by(code=data["campus_code"]).first()
            target_type = ProgramType.query.filter_by(name=str(data["program_type"]).upper()).first()
            target_year = AcademicYear.query.filter_by(name=data["academic_year"]).first()
            unit = OperatingUnit.query.filter_by(code=data.get("unit_code") or str(data["program_type"]).upper()).first()
            wing = Wing.query.filter_by(operating_unit_id=unit.id, code=data.get("wing_code")).first() if unit and data.get("wing_code") else None
            if not all((target_campus, target_type, target_year, unit)):
                raise ValueError(f"Project row {row.source_row} references unknown controlled data.")
            if not target:
                target = Project(code=data["code"], title=data["title"], campus_id=target_campus.id, program_type_id=target_type.id, academic_year_id=target_year.id, operating_unit_id=unit.id, wing_id=getattr(wing, "id", None), project_type=data.get("project_type") or "Operational", category=data.get("category") or "Operational", status="Draft", start_date=_parse_date(data["start_date"]), end_date=_parse_date(data["end_date"]))
                db.session.add(target)
        elif batch.import_type == "icc_roster":
            person = _person_for_import(data)
            target_year = AcademicYear.query.filter_by(name=data["academic_year"]).first()
            wing = Wing.query.filter_by(operating_unit_id=icc.id, code=data["wing_code"]).first()
            if not person or not target_year or not wing:
                raise ValueError(f"Roster row {row.source_row} has unresolved person, year, or wing.")
            target = TeamAssignment.query.filter_by(
                person_id=person.id,
                wing_id=wing.id,
                academic_year_id=target_year.id,
                project_id=None,
            ).first()
            if not target:
                target = TeamAssignment(
                    person_id=person.id,
                    wing_id=wing.id,
                    academic_year_id=target_year.id,
                    assignment_type="ICC Annual Roster",
                    role_label=data["role_label"],
                    recruitment_status="Selected",
                    starts_on=target_year.start_date,
                    ends_on=target_year.end_date,
                )
                db.session.add(target)
        else:
            project = Project.query.filter_by(code=data.get("project_code")).first()
            if not project:
                raise ValueError(f"Row {row.source_row} references an unknown project_code.")
            if batch.import_type == "participants":
                person = _person_for_import(data)
                if not person:
                    raise ValueError(f"Participant row {row.source_row} has no resolved person.")
                target = ProjectParticipant.query.filter_by(project_id=project.id, person_id=person.id).first()
                if not target:
                    target = ProjectParticipant(project_id=project.id, person_id=person.id, participant_type=data["participant_type"], status="Active")
                    db.session.add(target)
            elif batch.import_type == "attendance":
                person = _person_for_import(data)
                session = ProjectSession.query.filter_by(project_id=project.id, code=data["session_code"]).first()
                if not person or not session:
                    raise ValueError(f"Attendance row {row.source_row} has no resolved person or session.")
                target = SessionAttendance.query.filter_by(session_id=session.id, person_id=person.id).first()
                if not target:
                    target = SessionAttendance(session_id=session.id, person_id=person.id, status=data["status"], source_import_row_id=row.id)
                    db.session.add(target)
            elif batch.import_type == "checklists":
                template = ChecklistTemplate.query.filter_by(code=data["template_code"], is_active=True).order_by(ChecklistTemplate.version.desc()).first()
                if not template:
                    template = ChecklistTemplate(code=data["template_code"], name=data["template_code"], project_type=project.project_type or "Operational", version=1, source_reference=batch.source_file)
                    db.session.add(template)
                    db.session.flush()
                template_item = ChecklistTemplateItem.query.filter_by(template_id=template.id, code=data["item_code"]).first()
                if not template_item:
                    template_item = ChecklistTemplateItem(template_id=template.id, code=data["item_code"], title=data["title"], category=data.get("category"), sequence=sequence, mandatory=_as_bool(data.get("mandatory")), default_owner_label=data.get("owner"))
                    db.session.add(template_item)
                    db.session.flush()
                instance = ChecklistInstance.query.filter_by(project_id=project.id, template_id=template.id).first()
                if not instance:
                    instance = ChecklistInstance(project_id=project.id, template_id=template.id, name=f"{project.title} Checklist")
                    db.session.add(instance)
                    db.session.flush()
                target = ChecklistItemStatus.query.filter_by(checklist_instance_id=instance.id, template_item_id=template_item.id).first()
                if not target:
                    target = ChecklistItemStatus(checklist_instance_id=instance.id, template_item_id=template_item.id, external_owner=data.get("owner"), source_file=batch.source_file, source_sheet=row.sheet_name, source_row=row.source_row)
                    db.session.add(target)
            elif batch.import_type == "documents":
                target = DocumentRecord.query.filter_by(project_id=project.id, title=data["title"]).first()
                if not target:
                    classification = data.get("classification") or "Internal"
                    validation = validate_drive_link(data.get("drive_url"), classification) if data.get("drive_url") else {}
                    target = DocumentRecord(project_id=project.id, title=data["title"], category=data["category"], status=data.get("status") or "Missing", drive_url=data.get("drive_url") or None, drive_file_id=validation.get("file_id"), permission_classification=classification, mandatory_for_closure=_as_bool(data.get("mandatory_for_closure")))
                    db.session.add(target)
        db.session.flush()
        row.target_public_id = target.public_id
        committed += 1
    return committed


def commit_batch(batch, actor=None):
    # Re-fetch with a row lock so the staged->committed status check and the
    # eventual status write (below) are effectively one atomic transaction --
    # otherwise two concurrent commit_batch() calls could both pass the
    # "not yet committed" guard before either one flips the status, and both
    # proceed to write rows. On Postgres this is a real row lock held until
    # the transaction commits/rolls back; on SQLite with_for_update() is a
    # no-op (SQLite's own database-level write lock provides serialization
    # for the single-process dev/test case instead).
    locked_batch = ImportBatch.query.with_for_update().filter_by(id=batch.id).first()
    if locked_batch is not None:
        batch = locked_batch
    if batch.status == "Committed":
        return batch
    if batch.status not in {"Staged", "Validated"}:
        raise ValueError("Only a staged or validated import can be committed.")
    year, campus, icc_type, igp_type, icc, igp, events_wing = _reference_data()
    committed = 0
    valid_rows = ImportRow.query.filter_by(batch_id=batch.id, validation_status="Valid").order_by(ImportRow.id).all()

    if batch.import_type in STANDARD_IMPORTS:
        committed = _commit_standard_rows(batch, valid_rows, (year, campus, icc_type, igp_type, icc, igp, events_wing))

    elif batch.import_type == "events_summary":
        for sequence, row in enumerate(valid_rows, start=1):
            data = row.normalized_json
            event_date = _parse_date(data.get("event_date"))
            code = _code("ICC", sequence)
            project = Project.query.filter_by(code=code).first()
            if not project:
                project = Project(
                    code=code,
                    title=data["title"],
                    description="Imported from the 2026 ICC Events Report Summary.",
                    campus_id=campus.id,
                    program_type_id=icc_type.id,
                    academic_year_id=year.id,
                    operating_unit_id=icc.id,
                    wing_id=events_wing.id,
                    project_type="ICC event",
                    category="Event",
                    status="Planned",
                    start_date=event_date,
                    end_date=event_date,
                    venue=data.get("venue"),
                    target_audience=data.get("target_audience"),
                )
                db.session.add(project)
                db.session.flush()
                start_time, end_time = _parse_time_range(data.get("time_range"))
                db.session.add(ProjectSession(
                    project_id=project.id,
                    code="MAIN",
                    title=data["title"],
                    session_type="Event",
                    starts_at=datetime.combine(event_date, start_time, tzinfo=IST),
                    ends_at=datetime.combine(event_date, end_time, tzinfo=IST),
                    venue=data.get("venue"),
                ))
                for document in data.get("documents", []):
                    url = document["url"]
                    result = validate_drive_link(url) if "google.com" in url else {"file_id": None}
                    db.session.add(DocumentRecord(
                        project_id=project.id,
                        category=document["category"],
                        title=document["category"],
                        status="Submitted",
                        drive_url=url,
                        drive_file_id=result.get("file_id"),
                        permission_classification="Internal",
                    ))
            row.target_public_id = project.public_id
            committed += 1

    elif batch.import_type == "coffee_meet":
        project = Project.query.filter_by(code="ICC-2026-CEN-001").first()
        if not project:
            raise ValueError("Commit the events summary before Coffee Meet & Greet details.")
        for row in valid_rows:
            data = row.normalized_json
            if row.target_entity == "Person":
                registration = data.get("registration_number")
                person = Person.query.filter_by(registration_number=registration).first() if registration else None
                if not person:
                    parts = data["name"].split(maxsplit=1)
                    person = Person(
                        registration_number=registration,
                        first_name=parts[0],
                        last_name=parts[1] if len(parts) > 1 else None,
                        campus_id=campus.id,
                        person_type="Student",
                        nationality_country=data.get("category"),
                    )
                    db.session.add(person)
                    db.session.flush()
                assignment = TeamAssignment.query.filter_by(person_id=person.id, project_id=project.id).first()
                if not assignment:
                    db.session.add(TeamAssignment(
                        person_id=person.id,
                        project_id=project.id,
                        assignment_type="Project Team",
                        role_label=data.get("role") or "Volunteer",
                        recruitment_status="Selected",
                        starts_on=project.start_date,
                        ends_on=project.end_date,
                    ))
                row.target_public_id = person.public_id
            elif row.target_entity == "WorkTask":
                task = WorkTask.query.filter_by(project_id=project.id, title=data["title"]).first()
                if not task:
                    task = WorkTask(
                        project_id=project.id,
                        title=data["title"],
                        description=data.get("details"),
                        status="Completed" if re.search(r"confirmed|shared|complete", data.get("details") or "", re.I) else "In Progress",
                        mandatory_for_closure=True,
                    )
                    db.session.add(task)
                    db.session.flush()
                row.target_public_id = task.public_id
            elif row.target_entity == "DocumentRecord":
                document = DocumentRecord.query.filter_by(project_id=project.id, title=data["title"]).first()
                if not document:
                    document = DocumentRecord(
                        project_id=project.id,
                        category="Source Folder",
                        title=data["title"],
                        status="Indexed",
                        permission_classification="Internal",
                    )
                    db.session.add(document)
                    db.session.flush()
                row.target_public_id = document.public_id
            elif row.target_entity == "ProjectSession":
                session = ProjectSession.query.filter_by(project_id=project.id, code=data["code"]).first()
                if not session:
                    starts = datetime.strptime(f"2026-06-05 {data['starts']}", "%Y-%m-%d %H:%M").replace(tzinfo=IST)
                    ends = datetime.strptime(f"2026-06-05 {data['ends']}", "%Y-%m-%d %H:%M").replace(tzinfo=IST)
                    session = ProjectSession(
                        project_id=project.id,
                        code=data["code"],
                        title=data["title"],
                        session_type="Programme item",
                        starts_at=starts,
                        ends_at=ends,
                        venue=project.venue,
                    )
                    db.session.add(session)
                    db.session.flush()
                row.target_public_id = session.public_id
            elif row.target_entity == "ReportSnapshot":
                report = ReportSnapshot.query.filter_by(project_id=project.id, report_type="ICC Event Narrative").first()
                if not report:
                    report = ReportSnapshot(
                        project_id=project.id,
                        report_type="ICC Event Narrative",
                        title=data["title"],
                        snapshot_json={
                            "narrative": data["narrative"],
                            "actual_reach": data["actual_reach"],
                            "audience_breakdown": data["audience_breakdown"],
                        },
                        source_references=[row.public_id, batch.public_id],
                        approval_status="Draft",
                    )
                    db.session.add(report)
                    db.session.flush()
                project.actual_reach = data["actual_reach"]
                project.closure_summary = data["narrative"]
                row.target_public_id = report.public_id
            committed += 1

    elif batch.import_type == "summer_school":
        project = Project.query.filter_by(code="IGP-2026-CEN-001").first()
        if not project:
            project = Project(
                code="IGP-2026-CEN-001",
                title="International Summer School 2026",
                description="IGP program created from the supplied Summer School checklist.",
                campus_id=campus.id,
                program_type_id=igp_type.id,
                academic_year_id=year.id,
                operating_unit_id=igp.id,
                project_type="IGP inbound program",
                category="Summer School",
                status="Planned",
                start_date=date(2026, 6, 29),
                end_date=date(2026, 7, 31),
            )
            db.session.add(project)
            db.session.flush()
        template = ChecklistTemplate.query.filter_by(code="IGP-SUMMER-SCHOOL", version=1).first()
        if not template:
            template = ChecklistTemplate(
                code="IGP-SUMMER-SCHOOL",
                name="IGP Summer School Operational Checklist",
                project_type="IGP inbound program",
                version=1,
                source_reference=batch.source_file,
            )
            db.session.add(template)
            db.session.flush()
        instance = ChecklistInstance.query.filter_by(project_id=project.id, template_id=template.id).first()
        if not instance:
            instance = ChecklistInstance(project_id=project.id, template_id=template.id, name="Summer School 2026 Checklist")
            db.session.add(instance)
            db.session.flush()
        for sequence, row in enumerate(valid_rows, start=1):
            data = row.normalized_json
            item = ChecklistTemplateItem.query.filter_by(template_id=template.id, code=data["code"]).first()
            if not item:
                item = ChecklistTemplateItem(
                    template_id=template.id,
                    code=data["code"],
                    title=data["title"],
                    category=data.get("category"),
                    sequence=sequence,
                    mandatory=data.get("mandatory", True),
                    default_owner_label=data.get("external_owner"),
                    sensitive=data.get("sensitive", False),
                )
                db.session.add(item)
                db.session.flush()
            status = ChecklistItemStatus.query.filter_by(checklist_instance_id=instance.id, template_item_id=item.id).first()
            if not status:
                status = ChecklistItemStatus(
                    checklist_instance_id=instance.id,
                    template_item_id=item.id,
                    external_owner=data.get("external_owner"),
                    status="Not Started",
                    decision_comment=data.get("source_status"),
                    source_file=batch.source_file,
                    source_sheet=row.sheet_name,
                    source_row=row.source_row,
                )
                db.session.add(status)
                db.session.flush()
            row.target_public_id = status.public_id
            committed += 1

    batch.status = "Committed"
    batch.committed_count = committed
    batch.committed_at = db.func.now()
    batch.committed_by_id = getattr(actor, "id", None)
    batch.reconciliation_json = {
        "staged": batch.staged_count,
        "valid": batch.valid_count,
        "errors": batch.error_count,
        "committed": committed,
        "difference": batch.valid_count - committed,
    }
    record_audit("import.commit", batch, after=batch.reconciliation_json, actor=actor)
    db.session.commit()
    return batch


def seed_icc_checklist_template():
    """The only checklist template that exists (IGP-SUMMER-SCHOOL) came from
    the supplied Summer School source; there is no ICC-side equivalent, so
    ICC wing leaders had nothing to pick in the checklist step. Seeds a
    small, generic ICC event checklist template. Idempotent."""
    template = ChecklistTemplate.query.filter_by(code="ICC-EVENT-STANDARD", is_active=True).first()
    if template:
        return {"created": False}
    template = ChecklistTemplate(code="ICC-EVENT-STANDARD", name="ICC Event Standard Checklist", project_type="ICC event", version=1, source_reference="Manual seed")
    db.session.add(template)
    db.session.flush()
    items = [
        ("VENUE", "Venue booking confirmed", "Logistics"),
        ("POSTER", "Poster / invite designed", "Media"),
        ("ROSTER", "Volunteer roster finalized", "People"),
        ("ATTENDANCE", "Attendance sheet prepared", "Operations"),
        ("REPORT", "Post-event report submitted", "Reporting"),
    ]
    for sequence, (code, title, category) in enumerate(items, start=1):
        db.session.add(ChecklistTemplateItem(template_id=template.id, code=code, title=title, category=category, sequence=sequence, mandatory=True))
    db.session.commit()
    return {"created": True, "items": len(items)}


def backfill_summer_school_sample_content():
    """The supplied Summer School source (`_Summer School- Check List.xlsx`)
    is a checklist only -- it carries no schedule, roster, or document data,
    so the imported project has zero sessions/team/documents even after a
    full commit. This adds a small, clearly schematic set of sessions/team/
    documents (reusing existing demonstrator Person records rather than
    inventing new named individuals) so the project is as complete a
    reference example as Coffee Meet & Greet. Idempotent: safe to re-run.
    """
    project = Project.query.filter_by(code="IGP-2026-CEN-001").first()
    if not project:
        return {"skipped": "IGP-2026-CEN-001 project not found; run demo-import-supplied first."}

    sessions_created = 0
    session_plan = [
        ("IGP-SS-ORIENT", "Orientation & Welcome", "Orientation", 0, 9, 13),
        ("IGP-SS-IMMERSION", "Cultural Immersion Week", "Workshop Series", 7, 9, 17),
        ("IGP-SS-EXCURSION", "Campus & City Excursion", "Excursion", 21, 9, 18),
        ("IGP-SS-CLOSING", "Closing Ceremony", "Ceremony", 32, 15, 18),
    ]
    for code, title, session_type, day_offset, start_hour, end_hour in session_plan:
        if ProjectSession.query.filter_by(project_id=project.id, code=code).first():
            continue
        session_date = project.start_date + timedelta(days=day_offset)
        db.session.add(ProjectSession(
            project_id=project.id, code=code, title=title, session_type=session_type,
            starts_at=datetime.combine(session_date, time(start_hour, 0)),
            ends_at=datetime.combine(session_date, time(end_hour, 0)),
            venue="Bangalore Central Campus",
        ))
        sessions_created += 1

    team_created = 0
    team_plan = [("IGP Program Team", "IGP Program Lead"), ("IGP Program Team", "Buddy Coordinator")]
    available_people = Person.query.order_by(Person.id).limit(len(team_plan)).all()
    for (assignment_type, role_label), person in zip(team_plan, available_people):
        if TeamAssignment.query.filter_by(project_id=project.id, person_id=person.id, assignment_type=assignment_type).first():
            continue
        db.session.add(TeamAssignment(person_id=person.id, project_id=project.id, assignment_type=assignment_type, role_label=role_label))
        team_created += 1

    documents_created = 0
    document_plan = [
        ("IGP Schedule / Itinerary", "Schedule"),
        ("Participant Handbook", "Report"),
    ]
    for title, category in document_plan:
        if DocumentRecord.query.filter_by(project_id=project.id, title=title).first():
            continue
        db.session.add(DocumentRecord(project_id=project.id, title=title, category=category, status="Missing", mandatory_for_closure=True))
        documents_created += 1

    db.session.commit()
    return {"sessions_created": sessions_created, "team_created": team_created, "documents_created": documents_created}
