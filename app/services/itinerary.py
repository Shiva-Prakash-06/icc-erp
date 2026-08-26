"""Itinerary bulk-import parsing and commit logic.

Two source layouts are supported:

- the *wide* layout (one row per date, one column per time slot plus
  Breakfast/Lunch/Dinner columns), as supplied in ``references/Summer School
  Itinerary`` -- an extensionless CSV; and
- the *canonical* layout (`Date, Start Time, End Time, Activity, Venue, Type,
  Breakfast, Lunch, Dinner`), one row per session.

Parsing is pure and side-effect free (`parse_itinerary`); `stage_itinerary_import`
and `commit_itinerary_batch` do the staging/commit against the database,
following the same `ImportBatch`/`ImportRow` provenance pattern used
elsewhere in `app.services.imports`.
"""

from __future__ import annotations

import csv
import hashlib
import io
import re
from datetime import date as date_cls, datetime, time, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from dateutil import parser as date_parser
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from app.database import db
from app.models.erp import ImportBatch, ImportRow, ItineraryRevision, ProjectSession
from app.services.audit import record_audit

IST = ZoneInfo("Asia/Kolkata")
IMPORTER_VERSION = "1"

MEAL_COLUMNS = ("breakfast", "lunch", "dinner")
CANONICAL_COLUMNS = ("date", "start time", "end time", "activity", "venue", "type", "breakfast", "lunch", "dinner")
IGNORED_CELL_VALUES = {"", "-", "--", "—"}

# A slot's activity is treated as spanning the whole day (no fabricated
# start/end time) when it is the only occupied slot for that date and its
# text matches one of these patterns. See PLAN.md "IGP itinerary".
ALL_DAY_PATTERNS = [
    re.compile(pattern, re.IGNORECASE)
    for pattern in (
        r"\barrival",
        r"\bdepart",
        r"\bholiday\b",
        r"\bexcursion",
        r"\bfree time\b",
        r"\bsightseeing\b",
        r"\brural exposure\b",
        r"\btravel\b",
        r"^return to\b",
    )
]

TIME_RANGE_RE = re.compile(
    r"(\d{1,2}(?::\d{2})?\s*(?:am|pm|noon|midnight))\s*[-–—]\s*(\d{1,2}(?::\d{2})?\s*(?:am|pm|noon|midnight))",
    re.IGNORECASE,
)
VENUE_RE = re.compile(r"\bvenue\b\s*:-?\s*([^\n|]+)", re.IGNORECASE)


class ItineraryParseError(ValueError):
    pass


def _clean(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "").strip()
    # The supplied itinerary CSV has a handful of cells with an unbalanced
    # stray quote character from broken source escaping (e.g. a cell that
    # opens with `"` but is never closed). Strip a lone leading/trailing
    # quote rather than surfacing it as literal activity text.
    if text.startswith('"') and not text.endswith('"'):
        text = text[1:].strip()
    elif text.endswith('"') and not text.startswith('"'):
        text = text[:-1].strip()
    return text


def _normalize(value: str) -> str:
    return re.sub(r"\s+", " ", _clean(value)).strip().lower()


def _looks_like_xlsx(content: bytes) -> bool:
    return content[:2] == b"PK"


def sniff_itinerary_format(filename: str, content: bytes) -> str:
    """Return 'xlsx' or 'csv' for an uploaded itinerary file.

    Accepts the extensionless supplied file by sniffing content when the
    filename has no/unknown suffix, rather than rejecting it outright.
    """
    suffix = Path(secure_filename(filename or "")).suffix.lower()
    if suffix == ".xlsx" or (not suffix and _looks_like_xlsx(content)):
        return "xlsx"
    if suffix == ".csv" or not suffix:
        try:
            content.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise ItineraryParseError("Unable to decode file as CSV text.")
        return "csv"
    raise ItineraryParseError(f"Unsupported itinerary file type: {suffix or '(none)'}")


def _rows_from_content(filename: str, content: bytes) -> list[list[str]]:
    file_format = sniff_itinerary_format(filename, content)
    if file_format == "xlsx":
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet = workbook.active
        return [[_clean(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return [[_clean(cell) for cell in row] for row in reader]


def _parse_time_token(token: str) -> time | None:
    token = token.strip().lower()
    if not token:
        return None
    if token == "noon":
        return time(12, 0)
    if token == "midnight":
        return time(0, 0)
    match = re.match(r"(\d{1,2})(?::(\d{2}))?\s*(am|pm)?", token)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2) or 0)
    meridiem = match.group(3)
    if meridiem == "pm" and hour != 12:
        hour += 12
    if meridiem == "am" and hour == 12:
        hour = 0
    if hour > 23:
        return None
    return time(hour, minute)


def _extract_time_range(text: str) -> tuple[time, time] | None:
    match = TIME_RANGE_RE.search(text)
    if not match:
        return None
    start = _parse_time_token(match.group(1))
    end = _parse_time_token(match.group(2))
    if start is None or end is None:
        return None
    return start, end


def _parse_date_cell(text: str) -> date_cls | None:
    text = _clean(text)
    if not text:
        return None
    try:
        parsed = date_parser.parse(text, fuzzy=True, dayfirst=False, default=datetime(2000, 1, 1))
    except (ValueError, OverflowError):
        return None
    return parsed.date()


def _is_all_day(activity_text: str) -> bool:
    return any(pattern.search(activity_text) for pattern in ALL_DAY_PATTERNS)


def _session_type(activity_text: str) -> str:
    lowered = activity_text.lower()
    if "excursion" in lowered or "sightseeing" in lowered:
        return "Excursion"
    if "lunch break" in lowered or lowered.strip() == "break":
        return "Break"
    if "holiday" in lowered:
        return "Holiday"
    if "arrival" in lowered:
        return "Arrival"
    if "depart" in lowered:
        return "Departure"
    return "Session"


def _find_title_row(rows: list[list[str]], header_row_idx: int) -> str | None:
    for row in rows[:header_row_idx]:
        non_blank = [cell for cell in row if _clean(cell)]
        if len(non_blank) == 1:
            return _clean(non_blank[0])
    return None


def _find_header_row(rows: list[list[str]]) -> int:
    for idx, row in enumerate(rows):
        if any(_normalize(cell) == "date" for cell in row):
            return idx
    raise ItineraryParseError("No header row containing a 'Date' column was found.")


def _is_canonical(header: list[str]) -> bool:
    normalized = {_normalize(cell) for cell in header if _clean(cell)}
    required = {"date", "start time", "end time", "activity"}
    return required.issubset(normalized)


def parse_itinerary(filename: str, content: bytes) -> dict:
    """Parse itinerary bytes into a structured, side-effect-free representation.

    Returns a dict with keys: title, start_date, end_date, default_venue,
    footer_notes (list[str]), warnings (list[str]), days (list of {date,
    sessions: [...], meals: {...}}).
    """
    rows = _rows_from_content(filename, content)
    header_row_idx = _find_header_row(rows)
    header = rows[header_row_idx]
    title = _find_title_row(rows, header_row_idx)
    if _is_canonical(header):
        return _parse_canonical(rows, header_row_idx, header, title)
    return _parse_wide(rows, header_row_idx, header, title)


def _parse_wide(rows, header_row_idx, header, title) -> dict:
    date_col = next(idx for idx, cell in enumerate(header) if _normalize(cell) == "date")
    slot_columns = []  # list of (col_idx, start_time, end_time)
    meal_columns = {}  # normalized meal name -> col_idx
    for idx in range(date_col + 1, len(header)):
        label = header[idx]
        normalized = _normalize(label)
        if not normalized:
            continue
        if normalized in MEAL_COLUMNS:
            meal_columns[normalized] = idx
            continue
        time_range = _extract_time_range(label)
        if time_range:
            slot_columns.append((idx, time_range[0], time_range[1]))

    footer_notes: list[str] = []
    warnings: list[str] = []
    days = []
    for row in rows[header_row_idx + 1:]:
        if len(row) <= date_col:
            continue
        parsed_date = _parse_date_cell(row[date_col])
        non_blank = [cell for cell in row if _clean(cell)]
        if parsed_date is None:
            if non_blank:
                footer_notes.append(" | ".join(_clean(cell) for cell in non_blank))
            continue

        occupied = []
        for idx, start_time, end_time in slot_columns:
            text = row[idx] if idx < len(row) else ""
            text = _clean(text)
            if text in IGNORED_CELL_VALUES:
                continue
            occupied.append({
                "col": idx,
                "text": text,
                "normalized": _normalize(text),
                "column_start": start_time,
                "column_end": end_time,
            })

        sessions = []
        group: list[dict] = []

        def flush_group():
            if not group:
                return
            first, last = group[0], group[-1]
            override = _extract_time_range(first["text"])
            start_time = override[0] if override else first["column_start"]
            end_override = _extract_time_range(last["text"])
            end_time = end_override[1] if end_override else last["column_end"]
            activity_text = first["text"]
            all_day = len(occupied) == 1 and _is_all_day(activity_text)
            sessions.append({
                "title": activity_text,
                "start_time": None if all_day else start_time,
                "end_time": None if all_day else end_time,
                "is_all_day": all_day,
                "session_type": _session_type(activity_text),
            })

        for entry in occupied:
            if group and group[-1]["normalized"] == entry["normalized"] and entry["col"] == group[-1]["col"] + 1:
                group.append(entry)
                continue
            flush_group()
            group = [entry]
        flush_group()

        meals = {}
        for meal_name, idx in meal_columns.items():
            value = row[idx] if idx < len(row) else ""
            meals[meal_name] = _clean(value) or None

        days.append({"date": parsed_date, "sessions": sessions, "meals": meals})

    if not days:
        raise ItineraryParseError("No dated itinerary rows were found.")

    default_venue = None
    for note in footer_notes:
        match = VENUE_RE.search(note)
        if match:
            default_venue = match.group(1).strip().rstrip(".,")
            break

    days.sort(key=lambda day: day["date"])
    return {
        "title": title,
        "start_date": days[0]["date"],
        "end_date": days[-1]["date"],
        "default_venue": default_venue,
        "footer_notes": footer_notes,
        "warnings": warnings,
        "days": days,
    }


def _parse_canonical(rows, header_row_idx, header, title) -> dict:
    normalized_header = [_normalize(cell) for cell in header]
    col = {name: normalized_header.index(name) for name in CANONICAL_COLUMNS if name in normalized_header}
    by_date: dict[date_cls, dict] = {}
    footer_notes: list[str] = []
    for row in rows[header_row_idx + 1:]:
        if not any(_clean(cell) for cell in row):
            continue
        raw_date = row[col["date"]] if col.get("date", -1) < len(row) else ""
        parsed_date = _parse_date_cell(raw_date)
        if parsed_date is None:
            footer_notes.append(" | ".join(_clean(cell) for cell in row if _clean(cell)))
            continue
        activity = _clean(row[col["activity"]]) if col.get("activity", -1) < len(row) else ""
        if activity in IGNORED_CELL_VALUES:
            continue
        start_time = _parse_time_token(_clean(row[col["start time"]])) if "start time" in col and col["start time"] < len(row) else None
        end_time = _parse_time_token(_clean(row[col["end time"]])) if "end time" in col and col["end time"] < len(row) else None
        override = _extract_time_range(activity)
        if override:
            start_time, end_time = override
        all_day = start_time is None or end_time is None
        session_type = _clean(row[col["type"]]) if "type" in col and col["type"] < len(row) else "" or _session_type(activity)
        venue = _clean(row[col["venue"]]) if "venue" in col and col["venue"] < len(row) else None
        day = by_date.setdefault(parsed_date, {"date": parsed_date, "sessions": [], "meals": {}})
        day["sessions"].append({
            "title": activity,
            "start_time": None if all_day else start_time,
            "end_time": None if all_day else end_time,
            "is_all_day": all_day,
            "session_type": session_type or _session_type(activity),
            "venue": venue,
        })
        for meal in MEAL_COLUMNS:
            if meal in col and col[meal] < len(row):
                value = _clean(row[col[meal]])
                if value:
                    day["meals"][meal] = value

    if not by_date:
        raise ItineraryParseError("No dated itinerary rows were found.")

    days = sorted(by_date.values(), key=lambda day: day["date"])
    return {
        "title": title,
        "start_date": days[0]["date"],
        "end_date": days[-1]["date"],
        "default_venue": None,
        "footer_notes": footer_notes,
        "warnings": [],
        "days": days,
    }


def _sequence_rows(parsed: dict):
    """Yield (source_key, day_date, session_dict, global_row_number)."""
    row_number = 0
    for day in parsed["days"]:
        for sequence, session in enumerate(day["sessions"]):
            source_key = f"{day['date'].isoformat()}#{sequence}"
            yield source_key, day["date"], session, row_number
            row_number += 1


def stage_itinerary_import(project, uploaded_file, operator_key):
    """Stage an itinerary file for a project. Idempotent per (project, file content)."""
    filename = secure_filename(uploaded_file.filename or "Summer School Itinerary")
    content = uploaded_file.read()
    digest = hashlib.sha256(content).hexdigest()
    key = f"v{IMPORTER_VERSION}:itinerary:{project.public_id}:{digest}"
    existing = ImportBatch.query.filter_by(idempotency_key=key).first()
    if existing:
        return existing

    parsed = parse_itinerary(filename, content)
    batch = ImportBatch(
        idempotency_key=key,
        import_type="itinerary",
        source_file=filename,
        source_sha256=digest,
        status="Staged",
        importer_version=IMPORTER_VERSION,
        project_id=project.id,
    )
    db.session.add(batch)
    db.session.flush()

    for source_key, day_date, session, row_number in _sequence_rows(parsed):
        messages = []
        if not session["title"]:
            messages.append("Activity text is required.")
        normalized = {
            "source_key": source_key,
            "date": day_date.isoformat(),
            "title": session["title"],
            "start_time": session["start_time"].isoformat() if session["start_time"] else None,
            "end_time": session["end_time"].isoformat() if session["end_time"] else None,
            "is_all_day": session["is_all_day"],
            "session_type": session["session_type"],
            "venue": session.get("venue"),
        }
        row = ImportRow(
            batch_id=batch.id,
            sheet_name="itinerary",
            source_row=row_number,
            source_json={"date": day_date.isoformat(), "source_key": source_key, "title": session["title"]},
            normalized_json=normalized,
            validation_status="Error" if messages else "Valid",
            validation_messages=messages,
            target_entity="ProjectSession",
        )
        db.session.add(row)
        batch.staged_count += 1
        batch.valid_count += 0 if messages else 1
        batch.error_count += 1 if messages else 0

    batch.reconciliation_json = {
        "title": parsed["title"],
        "start_date": parsed["start_date"].isoformat(),
        "end_date": parsed["end_date"].isoformat(),
        "default_venue": parsed["default_venue"],
        "footer_notes": parsed["footer_notes"],
        "day_count": len(parsed["days"]),
        "meals": {day["date"].isoformat(): day["meals"] for day in parsed["days"]},
    }
    record_audit("import.stage", batch, after={"type": "itinerary", "rows": batch.staged_count})
    db.session.commit()
    return batch


def create_igp_project_from_itinerary(uploaded_file, actor):
    """The 'Create IGP from itinerary' quick action: parses the file once to
    infer the program title and date envelope, creates a minimal IGP
    project, then stages and commits the same file as its itinerary."""
    from werkzeug.datastructures import FileStorage

    from app.services.project_quickcreate import create_minimal_project

    filename = secure_filename(uploaded_file.filename or "itinerary")
    content = uploaded_file.read()
    parsed = parse_itinerary(filename, content)
    project = create_minimal_project(
        program_type_name="IGP", title=parsed["title"] or "Untitled IGP Program",
        start_date=parsed["start_date"], end_date=parsed["end_date"], actor=actor,
        venue=parsed.get("default_venue"),
    )
    batch = stage_itinerary_import(project, FileStorage(stream=io.BytesIO(content), filename=filename), actor.public_id)
    commit_itinerary_batch(batch, actor)
    return project, batch


def commit_itinerary_batch(batch, actor):
    """Commit a staged itinerary batch: create a new active ItineraryRevision,
    upsert ProjectSession rows by stable source_key, and deactivate sessions
    from the prior revision that are no longer present (retaining any that
    carry attendance as historical records)."""
    if batch.import_type != "itinerary":
        raise ValueError("Not an itinerary batch.")
    if batch.status == "Committed":
        return batch
    locked = ImportBatch.query.filter_by(id=batch.id).with_for_update().one()
    if locked.status == "Committed":
        return locked

    project_id = locked.project_id
    meta = locked.reconciliation_json or {}
    revision = ItineraryRevision(
        project_id=project_id,
        import_batch_id=locked.id,
        source_document=locked.source_file,
        source_sha256=locked.source_sha256,
        parser_version=IMPORTER_VERSION,
        inferred_metadata=meta,
        warnings=[],
        is_active=True,
        created_by_id=getattr(actor, "id", None),
    )
    db.session.add(revision)
    db.session.flush()

    ItineraryRevision.query.filter(
        ItineraryRevision.project_id == project_id,
        ItineraryRevision.id != revision.id,
    ).update({"is_active": False})

    seen_keys = set()
    committed = 0
    for row in locked.rows:
        if row.validation_status != "Valid":
            continue
        data = row.normalized_json
        source_key = data["source_key"]
        seen_keys.add(source_key)
        day_date = date_cls.fromisoformat(data["date"])
        if data["is_all_day"]:
            starts_at = datetime.combine(day_date, time(0, 0), tzinfo=IST)
            ends_at = datetime.combine(day_date + timedelta(days=1), time(0, 0), tzinfo=IST)
        else:
            start_time = time.fromisoformat(data["start_time"]) if data["start_time"] else time(0, 0)
            end_time = time.fromisoformat(data["end_time"]) if data["end_time"] else time(23, 59)
            starts_at = datetime.combine(day_date, start_time, tzinfo=IST)
            ends_at = datetime.combine(day_date, end_time, tzinfo=IST)
            if ends_at < starts_at:
                ends_at = starts_at

        session = ProjectSession.query.filter_by(project_id=project_id, source_key=source_key).first()
        code = f"ITIN-{source_key}"
        if session is None:
            existing_by_code = ProjectSession.query.filter_by(project_id=project_id, code=code).first()
            session = existing_by_code or ProjectSession(project_id=project_id, code=code)
            db.session.add(session)
        session.title = data["title"][:180]
        session.session_type = data["session_type"] or "Session"
        session.starts_at = starts_at
        session.ends_at = ends_at
        session.venue = data.get("venue") or meta.get("default_venue")
        session.itinerary_revision_id = revision.id
        session.source_key = source_key
        session.is_all_day = data["is_all_day"]
        session.is_active = True
        session.import_batch_id = locked.id
        row.target_public_id = session.public_id
        committed += 1

    # Deactivate sessions from a previous itinerary revision that no longer
    # appear in this import. Sessions carrying attendance stay untouched
    # (retained as historical records) beyond having is_active cleared.
    stale = ProjectSession.query.filter(
        ProjectSession.project_id == project_id,
        ProjectSession.source_key.isnot(None),
        ProjectSession.source_key.notin_(seen_keys) if seen_keys else True,
        ProjectSession.itinerary_revision_id != revision.id,
    ).all()
    for session in stale:
        session.is_active = False

    locked.status = "Committed"
    locked.committed_count = committed
    locked.committed_at = db.func.now()
    locked.committed_by_id = getattr(actor, "id", None)
    locked.reconciliation_json = {
        **meta,
        "staged": locked.staged_count,
        "valid": locked.valid_count,
        "errors": locked.error_count,
        "committed": committed,
        "difference": locked.valid_count - committed,
        "deactivated": len(stale),
    }
    record_audit("import.commit", locked, after={"type": "itinerary", "committed": committed})
    db.session.commit()
    return locked
