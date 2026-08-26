"""Lightweight, project-scoped reimbursement entries: bulk import/export and
manual CRUD. Deliberately not a governance/payment-ledger workflow -- no
approval routing, just a status string. See PLAN.md "Reimbursements"."""

from __future__ import annotations

import csv
import hashlib
import io
from datetime import datetime
from decimal import Decimal, InvalidOperation

from dateutil import parser as date_parser
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

from app.database import db
from app.models.erp import ImportBatch, ImportRow, ReimbursementEntry
from app.services.audit import record_audit

IMPORTER_VERSION = "1"


class ReimbursementImportError(ValueError):
    pass


def _parse_date(value) -> datetime.date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    try:
        return date_parser.parse(text, fuzzy=True).date()
    except (ValueError, OverflowError):
        return None


def _parse_amount(value):
    if value is None or value == "":
        return None
    try:
        amount = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None
    return amount


def _read_rows(filename: str, content: bytes):
    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if suffix == "xlsx":
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        header = [str(cell or "").strip().lower() for cell in next(values)]
        for row in values:
            yield dict(zip(header, row))
    else:
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        for row in reader:
            yield {str(key).strip().lower(): value for key, value in row.items()}


def stage_reimbursement_import(project, uploaded_file, operator_key):
    filename = secure_filename(uploaded_file.filename or "reimbursements.csv")
    content = uploaded_file.read()
    digest = hashlib.sha256(content).hexdigest()
    key = f"v{IMPORTER_VERSION}:reimbursements:{project.public_id}:{digest}"
    existing = ImportBatch.query.filter_by(idempotency_key=key).first()
    if existing:
        return existing

    batch = ImportBatch(
        idempotency_key=key,
        import_type="reimbursements",
        source_file=filename,
        source_sha256=digest,
        status="Staged",
        importer_version=IMPORTER_VERSION,
        project_id=project.id,
    )
    db.session.add(batch)
    db.session.flush()

    for row_number, raw_row in enumerate(_read_rows(filename, content), start=2):
        if not any(str(value or "").strip() for value in raw_row.values()):
            continue
        raw_date = raw_row.get("date")
        raw_amount = raw_row.get("amount")
        party_name = str(raw_row.get("party name") or "").strip()
        bill_number = str(raw_row.get("bill number") or "").strip() or None
        particular = str(raw_row.get("particular") or "").strip() or None
        status = str(raw_row.get("status") or "").strip() or "Pending"

        parsed_date = _parse_date(raw_date)
        parsed_amount = _parse_amount(raw_amount)
        messages = []
        if not party_name:
            messages.append("Party Name is required.")
        if parsed_date is None:
            messages.append("Date is required and must be a valid date.")
        if parsed_amount is None:
            messages.append("Amount is required and must be a valid number.")
        elif parsed_amount < 0:
            messages.append("Amount cannot be negative.")

        row = ImportRow(
            batch_id=batch.id,
            sheet_name="reimbursements",
            source_row=row_number,
            source_json={
                "date": str(raw_date) if raw_date is not None else None,
                "party_name": party_name, "bill_number": bill_number,
                "amount": str(raw_amount) if raw_amount is not None else None,
                "particular": particular, "status": status,
            },
            normalized_json={
                "date": parsed_date.isoformat() if parsed_date else None,
                "party_name": party_name, "bill_number": bill_number,
                "amount": str(parsed_amount) if parsed_amount is not None else None,
                "particular": particular, "status": status,
            },
            validation_status="Error" if messages else "Valid",
            validation_messages=messages,
            target_entity="ReimbursementEntry",
        )
        db.session.add(row)
        batch.staged_count += 1
        batch.valid_count += 0 if messages else 1
        batch.error_count += 1 if messages else 0

    record_audit("import.stage", batch, after={"type": "reimbursements", "rows": batch.staged_count})
    db.session.commit()
    return batch


def commit_reimbursement_batch(batch, actor):
    if batch.import_type != "reimbursements":
        raise ValueError("Not a reimbursements batch.")
    if batch.status == "Committed":
        return batch
    locked = ImportBatch.query.filter_by(id=batch.id).with_for_update().one()
    if locked.status == "Committed":
        return locked

    committed = 0
    for row in locked.rows:
        if row.validation_status != "Valid":
            continue
        data = row.normalized_json
        entry = ReimbursementEntry(
            project_id=locked.project_id,
            date=datetime.fromisoformat(data["date"]).date(),
            party_name=data["party_name"],
            bill_number=data["bill_number"],
            amount=Decimal(data["amount"]),
            particular=data["particular"],
            status=data["status"],
            created_by_id=getattr(actor, "id", None),
            source_import_row_id=row.id,
        )
        db.session.add(entry)
        db.session.flush()
        row.target_public_id = entry.public_id
        committed += 1

    locked.status = "Committed"
    locked.committed_count = committed
    locked.committed_at = db.func.now()
    locked.committed_by_id = getattr(actor, "id", None)
    locked.reconciliation_json = {
        "staged": locked.staged_count,
        "valid": locked.valid_count,
        "errors": locked.error_count,
        "committed": committed,
        "difference": locked.valid_count - committed,
    }
    record_audit("import.commit", locked, after={"type": "reimbursements", "committed": committed})
    db.session.commit()
    return locked


def export_reimbursements(project, output_format="csv"):
    entries = ReimbursementEntry.query.filter_by(project_id=project.id).order_by(
        ReimbursementEntry.date, ReimbursementEntry.id
    ).all()
    columns = ["S. No", "Date", "Party Name", "Bill Number", "Amount", "Particular", "Status"]
    rows = [
        [index, entry.date.isoformat(), entry.party_name, entry.bill_number or "", str(entry.amount), entry.particular or "", entry.status]
        for index, entry in enumerate(entries, start=1)
    ]
    if output_format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reimbursements"
        sheet.append(columns)
        for row in rows:
            sheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    writer.writerows(rows)
    return io.BytesIO(output.getvalue().encode("utf-8"))
