"""Generalized, project-scoped ICC volunteer-attendance and event-folder
importers. Unlike `imports.py`'s `_stage_coffee` (a hardcoded demonstrator
fixture for one specific folder/project), these work for any ICC event.
See PLAN.md "ICC sources and attendance"."""

from __future__ import annotations

import io
import re

from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from app.database import db
from app.models.erp import DocumentRecord, ImportBatch, ImportRow, Person, ProjectSession, SessionAttendance, TeamAssignment
from app.services.audit import record_audit
from app.services.documents import (
    checksum_bytes,
    classify_filename,
    find_existing_by_checksum,
    find_replaceable,
    is_allowed_extension,
    supersede,
)
from app.services.people import find_duplicate_person

IMPORTER_VERSION = "1"

PRESENT_VALUES = {"present", "p", "yes", "1", "✓"}
ABSENT_VALUES = {"absent", "a", "no", "0"}

HEADER_ALIASES = {
    "s.no.": "serial", "s no": "serial", "sno": "serial", "s.no": "serial", "no.": "serial",
    "primary role": "role", "role": "role",
    "name": "name",
    "class": "programme", "programme": "programme", "program": "programme", "course": "programme",
    "reg number": "registration", "reg no.": "registration", "reg no": "registration",
    "registration": "registration", "registration number": "registration",
    "student type": "student_type", "category": "student_type", "type": "student_type",
    "attendance": "attendance",
    "parents name": "parent_name", "parent name": "parent_name",
}

GUEST_SHEET_NAMES = {"guest list", "guests"}


def map_attendance(value) -> str | None:
    text = str(value or "").strip().lower()
    if not text:
        return None
    if text in PRESENT_VALUES:
        return "Present"
    if text in ABSENT_VALUES:
        return "Absent"
    return None


def _find_header_row(rows: list[list]) -> int | None:
    for idx, row in enumerate(rows[:6]):
        normalized = {str(cell or "").strip().lower() for cell in row}
        if any(alias in normalized for alias in ("name",)):
            return idx
    return None


def _column_map(header_row: list) -> dict[str, int]:
    mapping = {}
    for idx, cell in enumerate(header_row):
        key = HEADER_ALIASES.get(str(cell or "").strip().lower())
        if key and key not in mapping:
            mapping[idx] = key
    return {value: key for key, value in mapping.items()}


def _cell(row, columns, name):
    idx = columns.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _ensure_main_session(project):
    session = ProjectSession.query.filter_by(project_id=project.id, code="MAIN").first()
    if session:
        return session
    from datetime import datetime, time
    from zoneinfo import ZoneInfo
    ist = ZoneInfo("Asia/Kolkata")
    session = ProjectSession(
        project_id=project.id, code="MAIN", title=project.title, session_type="Event",
        starts_at=datetime.combine(project.start_date, time(9, 0), tzinfo=ist),
        ends_at=datetime.combine(project.end_date, time(17, 0), tzinfo=ist),
    )
    db.session.add(session)
    db.session.flush()
    return session


def _roster_snapshot(project):
    roster = {}
    for assignment in TeamAssignment.query.filter_by(project_id=project.id).join(Person, TeamAssignment.person_id == Person.id).all():
        roster.setdefault(re.sub(r"\s+", " ", assignment.person.display_name.strip().lower()), []).append(assignment.person)
    return roster


def stage_icc_attendance_import(project, uploaded_file, operator_key):
    filename = secure_filename(uploaded_file.filename or "volunteer_attendance.xlsx")
    content = uploaded_file.read()
    digest = checksum_bytes(content)
    key = f"v{IMPORTER_VERSION}:icc_volunteer_attendance:{project.public_id}:{digest}"
    existing = ImportBatch.query.filter_by(idempotency_key=key).first()
    if existing:
        return existing

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    roster_sheet = workbook.worksheets[0]
    all_rows = [list(row) for row in roster_sheet.iter_rows(values_only=True)]
    header_idx = _find_header_row(all_rows)
    if header_idx is None:
        raise ValueError("No header row containing a Name column was found.")
    columns = _column_map(all_rows[header_idx])
    if "name" not in columns:
        raise ValueError("No Name column was found.")

    batch = ImportBatch(
        idempotency_key=key, import_type="icc_volunteer_attendance", source_file=filename,
        source_sha256=digest, status="Staged", importer_version=IMPORTER_VERSION, project_id=project.id,
    )
    db.session.add(batch)
    db.session.flush()

    row_number = header_idx + 1
    for row in all_rows[header_idx + 1:]:
        row_number += 1
        name = _cell(row, columns, "name")
        serial = _cell(row, columns, "serial")
        if not name:
            continue  # section/heading rows carry only a serial label
        registration = _cell(row, columns, "registration")
        if isinstance(registration, float) and registration.is_integer():
            registration = str(int(registration))
        elif registration is not None:
            registration = str(registration).strip()
        normalized = {
            "name": str(name).strip(),
            "registration_number": registration or None,
            "programme": str(_cell(row, columns, "programme") or "").strip() or None,
            "student_type": str(_cell(row, columns, "student_type") or "").strip() or None,
            "role": str(_cell(row, columns, "role") or "").strip() or None,
            "attendance": map_attendance(_cell(row, columns, "attendance")),
        }
        db.session.add(ImportRow(
            batch_id=batch.id, sheet_name=roster_sheet.title, source_row=row_number,
            source_json={"serial": serial, "name": name, "registration": registration},
            normalized_json=normalized,
            validation_status="Valid", validation_messages=[], target_entity="Person",
        ))
        batch.staged_count += 1
        batch.valid_count += 1

    # Guest sheet: only imported when it actually contains rows.
    guest_sheet = next((sheet for sheet in workbook.worksheets if sheet.title.strip().lower() in GUEST_SHEET_NAMES), None)
    guest_row_count = 0
    if guest_sheet is not None:
        guest_rows = [list(row) for row in guest_sheet.iter_rows(values_only=True)]
        guest_header_idx = _find_header_row(guest_rows)
        if guest_header_idx is not None:
            guest_columns = _column_map(guest_rows[guest_header_idx])
            gr = guest_header_idx + 1
            for row in guest_rows[guest_header_idx + 1:]:
                gr += 1
                name = _cell(row, guest_columns, "name")
                if not name:
                    continue
                guest_row_count += 1
                db.session.add(ImportRow(
                    batch_id=batch.id, sheet_name=guest_sheet.title, source_row=gr,
                    source_json={"name": name},
                    normalized_json={
                        "name": str(name).strip(),
                        "registration_number": str(_cell(row, guest_columns, "registration") or "").strip() or None,
                        "programme": str(_cell(row, guest_columns, "programme") or "").strip() or None,
                    },
                    validation_status="Valid", validation_messages=[], target_entity="Guest",
                ))
                batch.staged_count += 1
                batch.valid_count += 1

    batch.reconciliation_json = {"guest_rows": guest_row_count}
    record_audit("import.stage", batch, after={"type": "icc_volunteer_attendance", "rows": batch.staged_count})
    db.session.commit()
    return batch


def commit_icc_attendance_batch(batch, actor):
    if batch.import_type != "icc_volunteer_attendance":
        raise ValueError("Not an ICC volunteer-attendance batch.")
    if batch.status == "Committed":
        return batch
    locked = ImportBatch.query.filter_by(id=batch.id).with_for_update().one()
    if locked.status == "Committed":
        return locked

    from app.models.project import Project
    project = db.session.get(Project, locked.project_id)
    session = _ensure_main_session(project)
    roster = _roster_snapshot(project)

    committed = 0
    guest_present_count = 0
    for row in locked.rows:
        if row.validation_status != "Valid":
            continue
        data = row.normalized_json
        if row.target_entity == "Guest":
            guest_present_count += 1
            person = Person(first_name=data["name"].split()[0], last_name=" ".join(data["name"].split()[1:]) or None, person_type="Guest")
            db.session.add(person)
            db.session.flush()
            db.session.add(SessionAttendance(session_id=session.id, person_id=person.id, status="Present", source_import_row_id=row.id))
            row.target_public_id = person.public_id
            committed += 1
            continue

        person = find_duplicate_person(None, data.get("registration_number"))
        if person is None:
            normalized_name = re.sub(r"\s+", " ", data["name"].strip().lower())
            candidates = roster.get(normalized_name, [])
            person = candidates[0] if len(candidates) == 1 else None
        if person is None:
            name_parts = data["name"].split()
            person = Person(
                first_name=name_parts[0], last_name=" ".join(name_parts[1:]) or None,
                registration_number=data.get("registration_number"), person_type=data.get("student_type") or "Student",
            )
            db.session.add(person)
            db.session.flush()
        if not TeamAssignment.query.filter_by(project_id=project.id, person_id=person.id).first():
            db.session.add(TeamAssignment(
                person_id=person.id, project_id=project.id, assignment_type="Volunteer",
                role_label=data.get("role") or "Volunteer",
            ))

        if data["attendance"] is not None:
            existing = SessionAttendance.query.filter_by(session_id=session.id, person_id=person.id).first()
            if existing is None:
                db.session.add(SessionAttendance(
                    session_id=session.id, person_id=person.id, status=data["attendance"],
                    source_import_row_id=row.id,
                ))
        row.target_public_id = person.public_id
        committed += 1

    locked.status = "Committed"
    locked.committed_count = committed
    locked.committed_by_id = getattr(actor, "id", None)
    locked.committed_at = db.func.now()
    locked.reconciliation_json = {
        **(locked.reconciliation_json or {}),
        "staged": locked.staged_count, "valid": locked.valid_count,
        "committed": committed, "difference": locked.valid_count - committed,
    }
    inferred = infer_actual_reach(project, guest_present_count=guest_present_count)
    if project.actual_reach is None and inferred is not None:
        project.actual_reach = inferred
    record_audit("import.commit", locked, after={"type": "icc_volunteer_attendance", "committed": committed})
    db.session.commit()
    return locked


def infer_actual_reach(project, *, guest_present_count: int | None = None) -> int | None:
    """Precedence: explicitly labelled report total -> complete guest
    roster -> present session attendance -> blank. A volunteer-only roster
    is never treated as total event reach -- present attendance only counts
    when it belongs to a non-volunteer (participant/audience) assignment."""
    if project.actual_reach is not None:
        return project.actual_reach
    if guest_present_count:
        return guest_present_count
    session_ids = [row.id for row in ProjectSession.query.filter_by(project_id=project.id).with_entities(ProjectSession.id).all()]
    if not session_ids:
        return None
    non_volunteer_person_ids = {
        row.person_id for row in TeamAssignment.query.filter_by(project_id=project.id)
        .filter(TeamAssignment.assignment_type != "Volunteer")
        .with_entities(TeamAssignment.person_id).all()
    }
    if not non_volunteer_person_ids:
        return None
    present_count = SessionAttendance.query.filter(
        SessionAttendance.session_id.in_(session_ids),
        SessionAttendance.status == "Present",
        SessionAttendance.person_id.in_(non_volunteer_person_ids),
    ).count()
    return present_count or None


def stage_icc_event_folder_import(project, uploaded_files, actor):
    """Classify and attach a batch of uploaded event-folder files as
    DocumentRecords. Files matching an existing (title, category) are
    superseded automatically; identical checksums are deduplicated."""
    results = []
    total_bytes = 0
    for uploaded_file in uploaded_files:
        filename = secure_filename(uploaded_file.filename or "")
        if not filename:
            continue
        if not is_allowed_extension(filename):
            results.append({"filename": filename, "status": "rejected", "reason": "Unsupported file type."})
            continue
        content = uploaded_file.read()
        total_bytes += len(content)
        checksum = checksum_bytes(content)
        category, classification = classify_filename(filename)
        title = re.sub(r"\.[^.]+$", "", filename)

        existing = find_existing_by_checksum(project.id, checksum)
        if existing is not None:
            results.append({"filename": filename, "status": "duplicate", "document": existing.public_id})
            continue

        document = DocumentRecord(
            project_id=project.id, category=category, title=title, status="Indexed",
            permission_classification=classification, checksum_sha256=checksum,
            uploaded_by_id=getattr(actor, "id", None),
        )
        prior = find_replaceable(project.id, title, category)
        if prior is not None:
            supersede(prior, document)
        db.session.add(document)
        db.session.flush()
        results.append({"filename": filename, "status": "attached", "document": document.public_id, "category": category})

    record_audit("import.event_folder", project, after={"files": len(results)})
    db.session.commit()
    return results
