"""Buddy-allocation bulk importer.

Accepts the supplied `SNo, International Student, Christ Buddy, Contact`
layout (plus common header aliases), matches existing people by exact
normalized name within the project's own roster, and provisions minimal
`Person`/`User`/`RoleAssignment`/`TeamAssignment`/`BuddyAssignment` records
for anyone not already on the project. See PLAN.md "Buddy allocation".
"""

from __future__ import annotations

import csv
import hashlib
import io
import re

from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from app.database import db
from app.models.erp import ImportBatch, ImportRow, Person, TeamAssignment
from app.models.project import BuddyAssignment
from app.models.user import User
from app.services.audit import record_audit
from app.services.buddy import validate_buddy_assignment
from app.services.roles import replace_scoped_assignment

IMPORTER_VERSION = "1"

HEADER_ALIASES = {
    "sno": None,
    "s.no": None,
    "s no": None,
    "international student": "international_student",
    "international student name": "international_student",
    "exchange student": "international_student",
    "student": "international_student",
    "christ buddy": "buddy_name",
    "buddy": "buddy_name",
    "buddy name": "buddy_name",
    "contact": "contact",
    "contact number": "contact",
    "phone": "contact",
}


class BuddyImportError(ValueError):
    pass


def _normalize_name(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip()).lower()


def _split_name(full_name: str) -> tuple[str, str | None]:
    parts = (full_name or "").strip().split()
    if not parts:
        return "", None
    return parts[0], " ".join(parts[1:]) or None


def _read_rows(filename: str, content: bytes):
    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if suffix == "xlsx":
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        header = [str(cell or "").strip().lower() for cell in next(values)]
        for row in values:
            yield dict(zip(header, row))
    else:
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        for row in reader:
            yield {str(key).strip().lower(): value for key, value in row.items()}


def stage_buddy_import(project, uploaded_file, operator_key):
    filename = secure_filename(uploaded_file.filename or "buddy_allocation.csv")
    content = uploaded_file.read()
    digest = hashlib.sha256(content).hexdigest()
    key = f"v{IMPORTER_VERSION}:buddy_allocations:{project.public_id}:{digest}"
    existing = ImportBatch.query.filter_by(idempotency_key=key).first()
    if existing:
        return existing

    batch = ImportBatch(
        idempotency_key=key,
        import_type="buddy_allocations",
        source_file=filename,
        source_sha256=digest,
        status="Staged",
        importer_version=IMPORTER_VERSION,
        project_id=project.id,
    )
    db.session.add(batch)
    db.session.flush()

    for row_number, raw_row in enumerate(_read_rows(filename, content), start=2):
        normalized_row = {}
        for key, value in raw_row.items():
            mapped = HEADER_ALIASES.get(str(key).strip().lower())
            if mapped:
                normalized_row[mapped] = str(value).strip() if value is not None else ""
        if not any(normalized_row.values()):
            continue
        international_student = normalized_row.get("international_student", "")
        buddy_name = normalized_row.get("buddy_name", "")
        contact = normalized_row.get("contact", "")
        messages = []
        if not international_student:
            messages.append("International Student name is required.")
        if not buddy_name:
            messages.append("Christ Buddy name is required.")
        row = ImportRow(
            batch_id=batch.id,
            sheet_name="buddy_allocations",
            source_row=row_number,
            source_json={"international_student": international_student, "buddy_name": buddy_name, "contact": contact},
            normalized_json={
                "international_student": international_student,
                "buddy_name": buddy_name,
                "contact": contact,
            },
            validation_status="Error" if messages else "Valid",
            validation_messages=messages,
            target_entity="BuddyAssignment",
        )
        db.session.add(row)
        batch.staged_count += 1
        batch.valid_count += 0 if messages else 1
        batch.error_count += 1 if messages else 0

    record_audit("import.stage", batch, after={"type": "buddy_allocations", "rows": batch.staged_count})
    db.session.commit()
    return batch


def _roster_snapshot(project) -> dict[str, list[tuple[str, Person]]]:
    """Snapshot the project's team roster *before* this commit starts.

    Matching only against this pre-existing snapshot -- rather than
    re-querying live during the loop -- means two rows in the *same* file
    that happen to share a name (e.g. two different students both named
    "Georgia") are never silently merged into one person; only a name that
    already existed on the project before this import is treated as a
    candidate match (and re-imports still merge correctly against it).
    """
    roster: dict[str, list[tuple[str, Person]]] = {}
    assignments = (
        TeamAssignment.query.filter_by(project_id=project.id)
        .join(Person, TeamAssignment.person_id == Person.id)
        .all()
    )
    for assignment in assignments:
        roster.setdefault(assignment.role_label or "", []).append(
            (_normalize_name(assignment.person.display_name), assignment.person)
        )
    return roster


def _find_team_person(roster: dict, name: str, role_label: str) -> Person | None:
    normalized = _normalize_name(name)
    matches = [person for candidate_name, person in roster.get(role_label, []) if candidate_name == normalized]
    if len(matches) > 1:
        raise BuddyImportError(f"Multiple existing {role_label} records named '{name}' on this project; resolve manually.")
    return matches[0] if matches else None


def _generate_username(first_name: str, contact: str | None) -> str:
    base = re.sub(r"[^a-z0-9]", "", first_name.lower()) or "buddy"
    candidate = base
    if not User.query.filter_by(username=candidate).first():
        return candidate
    if contact:
        digits = re.sub(r"\D", "", contact)[-4:]
        if digits:
            candidate = f"{base}{digits}"
            if not User.query.filter_by(username=candidate).first():
                return candidate
    suffix = 2
    while User.query.filter_by(username=f"{base}{suffix}").first():
        suffix += 1
    return f"{base}{suffix}"


def commit_buddy_batch(batch, actor):
    if batch.import_type != "buddy_allocations":
        raise ValueError("Not a buddy-allocations batch.")
    if batch.status == "Committed":
        return batch
    locked = ImportBatch.query.filter_by(id=batch.id).with_for_update().one()
    if locked.status == "Committed":
        return locked

    from app.models.project import Project  # local import avoids a cycle at module load time
    project = db.session.get(Project, locked.project_id)
    default_password = _config_default_password()
    roster = _roster_snapshot(project)

    committed = 0
    for row in locked.rows:
        if row.validation_status != "Valid":
            continue
        data = row.normalized_json
        try:
            student_person = _find_team_person(roster, data["international_student"], "Exchange Student")
            if student_person is None:
                first, last = _split_name(data["international_student"])
                student_person = Person(first_name=first, last_name=last, person_type="Student")
                db.session.add(student_person)
                db.session.flush()
                db.session.add(TeamAssignment(
                    person_id=student_person.id, project_id=project.id,
                    assignment_type="Participant", role_label="Exchange Student",
                ))

            buddy_person = _find_team_person(roster, data["buddy_name"], "Buddy")
            buddy_user = buddy_person.user_account if buddy_person else None
            if buddy_person is None:
                if not default_password:
                    raise BuddyImportError(
                        "AUTO_PROVISIONED_BUDDY_DEFAULT_PASSWORD is not configured; cannot provision buddy accounts."
                    )
                first, last = _split_name(data["buddy_name"])
                buddy_person = Person(first_name=first, last_name=last, person_type="Student", phone=data.get("contact") or None)
                db.session.add(buddy_person)
                db.session.flush()
                username = _generate_username(first, data.get("contact"))
                buddy_user = User(
                    username=username, email=None, role="Buddy", preferred_role="Buddy",
                    status="Approved", needs_password_reset=True, person_id=buddy_person.id,
                )
                buddy_user.set_password(default_password)
                db.session.add(buddy_user)
                db.session.flush()
                replace_scoped_assignment(buddy_user, "Buddy", {"project_public_id": project.public_id}, actor)
                db.session.add(TeamAssignment(
                    person_id=buddy_person.id, project_id=project.id, user_id=buddy_user.id,
                    assignment_type="Buddy", role_label="Buddy",
                ))

            validate_buddy_assignment(
                project, buddy_user.id if buddy_user else None, None,
                project.start_date, project.end_date,
                buddy_person_id=None if buddy_user else buddy_person.id,
                exchange_student_person_id=student_person.id,
            )
            existing_assignment = BuddyAssignment.query.filter_by(
                project_id=project.id,
                buddy_user_id=buddy_user.id if buddy_user else None,
                exchange_student_person_id=student_person.id,
            ).first()
            if existing_assignment is None:
                assignment = BuddyAssignment(
                    project_id=project.id,
                    buddy_user_id=buddy_user.id if buddy_user else None,
                    buddy_person_id=None if buddy_user else buddy_person.id,
                    exchange_student_person_id=student_person.id,
                    start_date=project.start_date, end_date=project.end_date,
                    source_import_row_id=row.id,
                )
                db.session.add(assignment)
            row.target_public_id = buddy_person.public_id
            committed += 1
        except (BuddyImportError, ValueError) as error:
            row.validation_status = "Error"
            row.validation_messages = [str(error)]
            batch.error_count += 1
            batch.valid_count = max(0, batch.valid_count - 1)
            continue

    locked.status = "Committed"
    locked.committed_count = committed
    locked.committed_at = db.func.now()
    locked.committed_by_id = getattr(actor, "id", None)
    locked.reconciliation_json = {
        "staged": locked.staged_count,
        "valid": locked.valid_count,
        "errors": locked.error_count,
        "committed": committed,
        "difference": locked.valid_count - committed,
    }
    record_audit("import.commit", locked, after={"type": "buddy_allocations", "committed": committed})
    db.session.commit()
    return locked


def _config_default_password():
    from flask import current_app
    return current_app.config.get("AUTO_PROVISIONED_BUDDY_DEFAULT_PASSWORD")
